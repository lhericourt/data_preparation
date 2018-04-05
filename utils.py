# -*- coding: utf-8 -*-
import logging
import itertools as it

import pandas as pd
import numpy as np
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Set up python logging format
log_format = '%(asctime)s %(levelname)s %(message)s'
logging.basicConfig(format=log_format, level=logging.INFO)


class ExcelUtils(object):
    @staticmethod
    def create_excel_file(file_path, sheet_name, df):
        """
        Initialize an excel file from a dataframe
        :param file_path: file path to save the Excel file
        :param sheet_name: name of the sheet where we want to save the dataframe
        :param df: dataframe to save 
        :return: None
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        for row in dataframe_to_rows(df, index=True, header=True):
            ws.append(row)
        wb.save(filename=file_path)

    @staticmethod
    def add_excel_sheet(file_path, sheet_name, df):
        """
        Add a sheet from a dataframe into an existing Excel file
        :param file_path: file path of the Excel file to update
        :param sheet_name: new sheet to add into the Excel file
        :param df: dataframe to save in the the sheet
        :return: None
        """
        wb = load_workbook(filename=file_path)
        ws = wb.create_sheet(title=sheet_name)
        for row in dataframe_to_rows(df, index=True, header=True):
            ws.append(row)
        wb.save(filename=file_path)


class DBUtils(object):
    @staticmethod
    def get_distinct_values_from_db(attribute, user, password, dbname, host):
        """
        Get all different values for one attribute of the database
        :return: the different values in a list
        """
        try:
            conn = psycopg2.connect("dbname=" + dbname + " user=" + user + " host=" + host + " password=" + password)
        except:
            print("I am unable to connect to the database")
            return

        cur = conn.cursor()
        cur.execute("SELECT DISTINCT " + attribute + " FROM car.target_car")
        distinct_val = cur.fetchall()
        cur.close()

        distinct_val = [x[0] for x in distinct_val]

        return distinct_val

    @staticmethod
    def get_distinct_model_and_variant(user, password, dbname, host):
        """
        Get all different values of models and variant
        :return: the different values into a dataframe
        """
        try:
            conn = psycopg2.connect("dbname=" + dbname + " user=" + user + " host=" + host + " password=" + password)
        except:
            print("I am unable to connect to the database")
            return

        cur = conn.cursor()
        cur.execute("SELECT DISTINCT make, model, model_variant FROM car.target_car")
        distinct_val = cur.fetchall()
        cur.close()
        distinct_val_df = pd.DataFrame(distinct_val, columns=["make", "model", "model_variant"])
        return distinct_val_df


class BusinessRules(object):
    def __init__(self):
        self.color_mapping = {
            'anthrazit': 'Gray',
            'anthrazit mét.': 'Gray',
            'beige': 'Beige',
            'beige mét.': 'Beige',
            'blau': 'Blue',
            'blau mét.': 'Blue',
            'bordeaux': 'Other',
            'bordeaux mét.': 'Other',
            'braun': 'Brown',
            'braun mét.': 'Brown',
            'gelb': 'Yellow',
            'gelb mét.': 'Yellow',
            'gold': 'Gold',
            'gold mét.': 'Gold',
            'grau': 'Gray',
            'grau mét.': 'Gray',
            'grün': 'Green',
            'grün mét.': 'Green',
            'orange': 'Orange',
            'orange mét.': 'Orange',
            'rot': 'Red',
            'rot mét.': 'Red',
            'schwarz': 'Black',
            'schwarz mét.': 'Black',
            'silber': 'Gray',
            'silber mét.': 'Gray',
            'violett mét.': 'Purple',
            'weiss': 'White',
            'weiss mét.': 'White',
            np.NaN: np.NaN}

        self.type_mapping = {
            'Coupé': 'Coupé',
            'Limousine': 'Other',
            'Cabriolet': 'Convertible / Roadster',
            'Kombi': 'Other',
            'SUV / Geländewagen': 'SUV',
            'Kleinwagen': 'Other',
            'Kompaktvan / Minivan': 'Station Wagon',
            'Pick-up': 'Other',
            'Sattelschlepper': 'Other',
            'Wohnkabine': 'Other',
            np.NaN: np.NaN}

        self.condition_mapping = {
            'Occasion': 'Used',
            'Oldtimer': 'Used',
            'Neu': 'New',
            'Vorführmodell': 'New',
            np.NaN: np.NaN}

    @staticmethod
    def compute_car_type(car, type_mapping):
        """
        Compute the type of one car depending on the attributes properties, bodytypetext and seats
        :param car: the car to compute the type
        :param type_mapping: the dictionary defining the mapping between the type values of the source and the target
        :return: the type of the car
        """
        if isinstance(car["Properties"], str):
            if "tuning" in car["Properties"].lower():
                car_type = "Custom"
                return car_type

        if car["Seats"] == "1":
            car_type = "Single Seater"
            return car_type

        car_type = type_mapping[car["BodyTypeText"]]

        return car_type

    @staticmethod
    def compute_make(car, list_makes):
        """
        Compute the make of one car : it warns the user when the make is not known from the database
        :param car: the car to compute the make
        :param list_makes: list of makes known by the database
        :return: the make in the database format
        """
        list_makes_lower_case = [x.lower() for x in list_makes]
        if car["MakeText"].lower() in list_makes_lower_case:
            return list_makes[list_makes_lower_case.index(car["MakeText"].lower())]
        else:
            logging.warning("The maker {} of the car with the ID {} is not known "
                            "from the database".format(car["MakeText"], car["ID"]))
            return car["MakeText"]

    @staticmethod
    def compute_model(car, df_models_variants):
        """
        Compute the model of one car: it warn the user when the model is not known from the database
        :param car: the car to compute the model
        :param df_models_variants: list of models and model variants known by the database
        :return: the model in the database format
        """

        # We get the models and the variants for the make of the car
        df_models_variants_filtered = df_models_variants[df_models_variants["make"] == car["make"]]
        models = list(df_models_variants_filtered["model"])
        variants = list(df_models_variants_filtered["model_variant"])

        # We delete spaces and case of models and variants
        models_no_space = [x.lower().replace(" ", "") for x in models]
        variants_no_space = [x.lower().replace(" ", "") for x in variants]

        # We try all the permutation of the words of the model, and if one is contained in a model or a variant
        # of the database, then we return the corresponding model
        if str(car["ModelText"]) != "nan":
            list_word_permutation = list(it.permutations(car["ModelText"].split(" ")))
            for word_perm in list_word_permutation:
                word_perm_concat = "".join(word_perm).lower()
                for i, model_variant in enumerate(zip(models_no_space, variants_no_space)):
                    if word_perm_concat in model_variant[0] + model_variant[1]:
                        return models[i]

            logging.warning("The model {} of the car with the ID {} is not known from"
                            " the database".format(car["ModelText"], car["ID"]))
        return car["ModelText"]

    @staticmethod
    def compute_variant(car):
        """
        Compute the model variant of one car: to avoid useless information it removes the model name of the model 
        variant when the variant starts with it
        :param car: the car to compute the model variant
        :return: the variant
        """
        if str(car["model"]) != "nan":
            if car["ModelTypeText"].lower().startswith(car["model"].lower()):
                return car["ModelTypeText"][len(car["model"]):]

        return car["ModelTypeText"]
